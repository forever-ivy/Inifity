#!/usr/bin/env python3
"""One-off repair for truncated spreadsheet translations in a job Final.xlsx."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.openclaw_translation_orchestrator import (
    SOURCE_TRUNCATED_MARKER,
    _codex_generate,
)


DEFAULT_JOB_DIR = Path(
    "/Users/ivy/Library/CloudStorage/OneDrive-Personal/Translation Task/Translated -EN/_VERIFY/job_telegram_20260221_173907_361000be"
)
DEFAULT_SOURCE_XLSX = Path(
    "/Users/ivy/Library/CloudStorage/OneDrive-Personal/Translation Task/_INBOX/telegram/job_telegram_20260221_173907_361000be/FD .xlsx"
)


def _has_terminal_punctuation(text: str) -> bool:
    value = str(text or "").strip()
    if not value:
        return False
    if value.endswith(SOURCE_TRUNCATED_MARKER):
        return True
    return value[-1] in ".!?\"'):]}>"


def _cell_key(sheet: str, cell: str) -> str:
    return f"{sheet}!{cell.upper()}"


def _build_context(source_xlsx: Path, unit: dict[str, str]) -> dict[str, Any]:
    return _build_batch_context(source_xlsx, [unit])


def _build_batch_context(source_xlsx: Path, units: list[dict[str, str]]) -> dict[str, Any]:
    return {
        "subject": "Repair truncated spreadsheet translation cell-by-cell",
        "message_text": "Translate Arabic cell to complete English output without truncation.",
        "task_intent": {
            "task_type": "SPREADSHEET_TRANSLATION",
            "source_language": "ar",
            "target_language": "en",
        },
        "candidate_files": [
            {
                "path": str(source_xlsx),
                "name": source_xlsx.name,
                "language": "ar",
            }
        ],
        "format_preserve": {
            "xlsx_sources": [
                {
                    "file": source_xlsx.name,
                    "path": str(source_xlsx),
                    "cell_units": units,
                }
            ]
        },
    }


def _translate_batch(
    source_xlsx: Path,
    units: list[dict[str, str]],
    *,
    max_attempts: int = 2,
) -> tuple[bool, dict[str, str], str]:
    findings = [
        "xlsx_truncated_cell_repair",
        "CRITICAL: translate every provided cell completely from beginning to end.",
        "Do not stop mid-sentence. Do not summarize. Do not abbreviate.",
    ]
    for _ in range(max_attempts):
        out = _codex_generate(_build_batch_context(source_xlsx, units), None, findings, 1)
        if not out.get("ok"):
            continue
        data = out.get("data") if isinstance(out.get("data"), dict) else {}
        rows = data.get("xlsx_translation_map") if isinstance(data.get("xlsx_translation_map"), list) else []
        translated: dict[str, str] = {}
        for row in rows:
            if not isinstance(row, dict):
                continue
            sheet = str(row.get("sheet") or "").strip()
            cell = str(row.get("cell") or "").strip().upper()
            text = str(row.get("text") or "").strip()
            if sheet and cell and text:
                translated[_cell_key(sheet, cell)] = text
        if translated:
            return True, translated, ""
    return False, {}, "translation_failed"


def _translate_one_cell(
    source_xlsx: Path,
    unit: dict[str, str],
    *,
    max_attempts: int = 3,
) -> tuple[bool, str, str]:
    ok, translated, err = _translate_batch(source_xlsx, [unit], max_attempts=max_attempts)
    if not ok:
        return False, "", err
    key = _cell_key(unit["sheet"], unit["cell"])
    text = translated.get(key, "").strip()
    if text:
        return True, text, ""
    return False, "", "translation_failed"


def main() -> int:
    parser = argparse.ArgumentParser(description="Repair truncated xlsx translations for one job")
    parser.add_argument("--job-dir", default=str(DEFAULT_JOB_DIR))
    parser.add_argument("--source-xlsx", default=str(DEFAULT_SOURCE_XLSX))
    parser.add_argument("--sheet", default="Interview_FDs")
    parser.add_argument("--min-len", type=int, default=100)
    parser.add_argument("--batch-size", type=int, default=5)
    parser.add_argument("--max-attempts", type=int, default=2)
    args = parser.parse_args()

    job_dir = Path(args.job_dir).expanduser().resolve()
    source_xlsx = Path(args.source_xlsx).expanduser().resolve()
    final_xlsx = job_dir / "Final.xlsx"
    report_path = job_dir / ".system" / "xlsx_truncation_repair_report.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)

    wb_src = load_workbook(str(source_xlsx), data_only=False)
    wb_out = load_workbook(str(final_xlsx), data_only=False)
    ws_src = wb_src[args.sheet]
    ws_out = wb_out[args.sheet]

    candidates: list[dict[str, Any]] = []
    for row in range(1, ws_out.max_row + 1):
        cell = f"D{row}"
        src_val = ws_src[cell].value
        out_val = ws_out[cell].value
        if not isinstance(src_val, str) or not isinstance(out_val, str):
            continue
        src_text = src_val.strip()
        out_text = out_val.strip()
        if not src_text or not out_text:
            continue
        if len(out_text) <= args.min_len:
            continue
        if not _has_terminal_punctuation(out_text):
            candidates.append(
                {
                    "sheet": args.sheet,
                    "cell": cell,
                    "source_text": src_text,
                    "old_text": out_text,
                    "source_has_terminal": _has_terminal_punctuation(src_text),
                }
            )

    repaired: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []

    by_key = {_cell_key(item["sheet"], item["cell"]): item for item in candidates}
    keys = list(by_key.keys())
    for start in range(0, len(keys), max(1, args.batch_size)):
        batch_keys = keys[start : start + max(1, args.batch_size)]
        batch_units = []
        for key in batch_keys:
            item = by_key[key]
            batch_units.append(
                {
                    "file": source_xlsx.name,
                    "sheet": item["sheet"],
                    "cell": item["cell"],
                    "text": item["source_text"],
                }
            )
        ok, translated_map, _ = _translate_batch(
            source_xlsx,
            batch_units,
            max_attempts=max(1, args.max_attempts),
        )
        if not ok:
            translated_map = {}

        for key in batch_keys:
            item = by_key[key]
            new_text = str(translated_map.get(key) or "").strip()
            if not new_text:
                unit = {
                    "file": source_xlsx.name,
                    "sheet": item["sheet"],
                    "cell": item["cell"],
                    "text": item["source_text"],
                }
                single_ok, single_text, err = _translate_one_cell(
                    source_xlsx,
                    unit,
                    max_attempts=max(1, args.max_attempts),
                )
                if not single_ok:
                    failures.append({"cell": _cell_key(item["sheet"], item["cell"]), "error": err})
                    continue
                new_text = single_text
            marker_added = False
            if not item["source_has_terminal"] and not new_text.endswith(SOURCE_TRUNCATED_MARKER):
                new_text = f"{new_text.rstrip()} {SOURCE_TRUNCATED_MARKER}"
                marker_added = True
            if item["source_has_terminal"] and not _has_terminal_punctuation(new_text):
                failures.append(
                    {
                        "cell": _cell_key(item["sheet"], item["cell"]),
                        "error": "translated_text_still_truncated",
                    }
                )
                continue
            ws_out[item["cell"]] = new_text
            repaired.append(
                {
                    "cell": _cell_key(item["sheet"], item["cell"]),
                    "source_has_terminal": item["source_has_terminal"],
                    "old_len": len(item["old_text"]),
                    "new_len": len(new_text),
                    "marker_added": marker_added,
                }
            )

        print(f"processed_batch={start // max(1, args.batch_size) + 1} repaired={len(repaired)} failures={len(failures)}")

    wb_out.save(str(final_xlsx))
    wb_out.close()
    wb_src.close()

    report = {
        "job_dir": str(job_dir),
        "source_xlsx": str(source_xlsx),
        "final_xlsx": str(final_xlsx),
        "candidate_cells": len(candidates),
        "repaired_cells": len(repaired),
        "failed_cells": len(failures),
        "marker": SOURCE_TRUNCATED_MARKER,
        "repaired": repaired,
        "failures": failures,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
